"""
reporter.py  —  Generates the Servallab QC Excel workbook.

Sheets produced:
  1. SUMMARY              – dashboard: project info, KPIs, checks summary + bar chart
  2. ALL RECORDS          – every record with QC status columns appended
  3. FLAGGED RECORDS      – only flagged rows (all checks combined)
  4. <one sheet per check that produced flags>
  5. SUPPLEMENTAL <n>     – results pushed from individual tabs
  6. ITVR PERFORMANCE     – per-interviewer aggregated stats
  7. PRODUCTIVITY         – date × interviewer interview-count matrix
"""

import os
from datetime import datetime
from typing import List, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.utils import ensure_output_dir, setup_logger, timestamp_str
from core.validator import CheckResult

logger = setup_logger("reporter")


# ── Colour palette — 8-char ARGB (FF prefix = fully opaque) ──────────────────
NAVY   = "FF1F2B6C"
WHITE  = "FFFFFFFF"
DARK   = "FF333333"
ORANGE = "FFE65100"
GREEN  = "FF2E7D32"
RED    = "FFC62828"
AMBER  = "FFF9A825"
GRAY   = "FF757575"

FILL_NAVY       = PatternFill("solid", fgColor=NAVY)
FILL_RED        = PatternFill("solid", fgColor="FFFFEBEE")
FILL_ORANGE     = PatternFill("solid", fgColor="FFFFF3E0")
FILL_GREEN      = PatternFill("solid", fgColor="FFE8F5E9")
FILL_BLUE       = PatternFill("solid", fgColor="FFE3F2FD")
FILL_GRAY       = PatternFill("solid", fgColor="FFF5F5F5")
FILL_ALT        = PatternFill("solid", fgColor="FFF4F6FB")   # alternating row tint
FILL_AMBER_LITE = PatternFill("solid", fgColor="FFFFFDE7")

SEV_FILL  = {"critical": FILL_RED, "warning": FILL_ORANGE, "info": FILL_BLUE}
SEV_COLOR = {"critical": RED, "warning": ORANGE, "info": "FF1565C0"}

THIN = Side(style="thin", color="FFCCCCCC")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── Check → sheet name mapping ────────────────────────────────────────────────
_SHEET = {
    "missing_value_check":              "MISSING VALUES",
    "high_missing_column_check":        "MISSING COLUMNS",
    "range_check":                      "RANGE VIOLATIONS",
    "logic_check":                      "LOGIC VIOLATIONS",
    "pattern_check":                    "PATTERN VIOLATIONS",
    "duplicate_check":                  "DUPLICATES",
    "duration_check":                   "DURATION FLAGS",
    "anomaly_check":                    "STATISTICAL ANOMALIES",
    "straightlining_check":             "STRAIGHTLINING",
    "interviewer_duration_check":       "ITVR DURATION",
    "interviewer_productivity_check":   "ITVR PRODUCTIVITY",
    "consent_eligibility_check":        "CONSENT VIOLATIONS",
    "fabrication_check":                "FABRICATION FLAGS",
    "near_duplicate_check":             "NEAR DUPLICATES",
    "verbatim_quality_check":           "VERBATIM QUALITY",
}
_SEV_RANK = {"critical": 3, "warning": 2, "info": 1, "none": 0}

# ── Interviewer risk scoring (mirrors routers/interviewers.py) ────────────────
_RISK_WEIGHTS = {"fabrication": 0.40, "duration": 0.25, "straightlining": 0.25, "productivity": 0.10}
_RISK_FLAG_MAP = {
    "interviewer_duration_check":     "duration",
    "straightlining_check":           "straightlining",
    "fabrication_check":              "fabrication",
    "interviewer_productivity_check": "productivity",
}


# ── Low-level cell helpers ────────────────────────────────────────────────────

def _font(bold=False, size=11, color=DARK, italic=False, name="Calibri"):
    return Font(name=name, bold=bold, size=size, color=color, italic=italic)


def _safe(v):
    """Convert pandas NA types and lists to Excel-safe values."""
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(v, list):
        return ", ".join(str(x) for x in v)
    return v


def _cell(ws, row, col, value="", bold=False, size=11, color=DARK,
          fill=None, align="left", wrap=False, border=False, number_format=None):
    c = ws.cell(row=row, column=col, value=_safe(value))
    c.font = _font(bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill:
        c.fill = fill
    if border:
        c.border = THIN_BORDER
    if number_format:
        c.number_format = number_format
    if ws.row_dimensions[row].height in (None, 0, 15):
        ws.row_dimensions[row].height = 18
    return c


def _section_header(ws, row, col, text, span=5):
    c = _cell(ws, row, col, text, bold=True, size=11, color=WHITE, fill=FILL_NAVY)
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + span - 1)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             indent=1, wrap_text=False)
    ws.row_dimensions[row].height = 22
    return c


def _table_header(ws, row, col, headers, fill=FILL_NAVY, color=WHITE):
    for i, h in enumerate(headers):
        _cell(ws, row, col + i, h, bold=True, color=color,
              fill=fill, align="center", border=True)
    ws.row_dimensions[row].height = 20


def _auto_width(ws, min_w=8, max_w=55):
    for col in ws.columns:
        best = min_w
        for cell in col:
            if cell.value is not None:
                try:
                    w = len(str(cell.value))
                    if w > best:
                        best = w
                except Exception:
                    pass
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(best + 3, max_w)


def _freeze(ws, row=2, col=1):
    ws.freeze_panes = ws.cell(row=row, column=col)


# ── Reporter class ────────────────────────────────────────────────────────────

class Reporter:
    def __init__(self, output_dir: str = "outputs"):
        self.output_dir = output_dir
        ensure_output_dir(output_dir)

    # ── Public entry point ────────────────────────────────────────────────────

    def generate(self, results: List[CheckResult],
                 df_original: Optional[pd.DataFrame] = None,
                 supplemental: Optional[List[dict]] = None):
        """
        Generate the QC workbook.
        supplemental: list of dicts from ad-hoc tab runs, each with keys:
            check_name, issue_type, severity, flagged_rows (list of dicts)
        """
        ts = timestamp_str()
        wb = Workbook()
        wb.remove(wb.active)

        # Convert supplemental dicts to CheckResult objects
        extra_results = self._supplemental_to_results(supplemental or [])
        all_results = list(results) + extra_results

        self._write_summary(wb, all_results, df_original)
        self._write_all_records(wb, all_results, df_original)
        self._write_flagged_records(wb, all_results)
        for result in all_results:
            if result.flag_count > 0:
                self._write_check_sheet(wb, result, df_original)
        if df_original is not None:
            self._write_itvr_performance(wb, all_results, df_original)
            self._write_productivity(wb, df_original)
            self._write_risk_scorecard(wb, all_results, df_original)

        path = os.path.join(self.output_dir, f"qc_summary_{ts}.xlsx")
        wb.save(path)
        logger.info(f"QC report saved: {path}  ({len(wb.sheetnames)} sheets)")
        return path

    @staticmethod
    def _supplemental_to_results(supplemental: List[dict]) -> List[CheckResult]:
        extra = []
        seen_names: dict = {}
        for s in supplemental:
            base_name = s.get("check_name", "custom_check")
            seen_names[base_name] = seen_names.get(base_name, 0) + 1
            check_name = base_name if seen_names[base_name] == 1 else f"{base_name}_{seen_names[base_name]}"
            rows = pd.DataFrame(s.get("flagged_rows", []))
            extra.append(CheckResult(
                check_name=check_name,
                issue_type=s.get("issue_type", "Custom Check"),
                severity=s.get("severity", "warning"),
                flagged_rows=rows,
                metadata={},
            ))
        return extra

    # ── 1. SUMMARY ────────────────────────────────────────────────────────────

    def _write_summary(self, wb, results, df):
        ws = wb.create_sheet("SUMMARY")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 2
        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["C"].width = 26
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 4
        ws.column_dimensions["H"].width = 28
        ws.column_dimensions["I"].width = 18

        now = datetime.now().strftime("%d %B %Y, %H:%M")
        total_rows = len(df) if df is not None else 0
        total_flags = sum(r.flag_count for r in results)
        flagged_rows_count = self._count_unique_flagged_rows(results)
        error_rate = round(flagged_rows_count / max(total_rows, 1) * 100, 1)

        dur_col = self._detect_col(df, ["duration_minutes", "duration", "interview_duration"])
        avg_dur = round(df[dur_col].mean(), 1) if dur_col and df is not None else "—"

        # ── Title ─────────────────────────────────────────────────────────────
        ws.row_dimensions[1].height = 8
        ws.row_dimensions[2].height = 32
        _cell(ws, 2, 2, "SERVALLAB — QC ANALYSIS REPORT",
              bold=True, size=18, color=NAVY, align="left")
        _cell(ws, 3, 2, f"Generated: {now}", size=10, color="FF888888")
        ws.row_dimensions[4].height = 10

        # ── Project Info ──────────────────────────────────────────────────────
        r = 5
        _section_header(ws, r, 2, "PROJECT INFORMATION", span=4); r += 1
        info = [
            ("Dataset", df.attrs.get("filename", "Uploaded file") if df is not None else "—"),
            ("Total Records", total_rows),
            ("Checks Run", len(results)),
            ("Report Date", datetime.now().strftime("%Y-%m-%d")),
        ]
        for label, val in info:
            _cell(ws, r, 2, label, bold=True, color=NAVY, fill=FILL_ALT if r % 2 == 0 else None)
            _cell(ws, r, 3, val, color=DARK, fill=FILL_ALT if r % 2 == 0 else None)
            r += 1
        ws.row_dimensions[r].height = 10; r += 1

        # ── KPI Table ─────────────────────────────────────────────────────────
        _section_header(ws, r, 2, "KEY METRICS", span=4); r += 1
        _table_header(ws, r, 2, ["Metric", "Value"]); r += 1
        kpis = [
            ("Total Records",            total_rows,          "#,##0"),
            ("Total QC Flags",           total_flags,         "#,##0"),
            ("Unique Flagged Records",   flagged_rows_count,  "#,##0"),
            ("Error Rate",               error_rate / 100,    "0.0%"),
            ("Average Duration (min)",   avg_dur,             "0.0"),
            ("Critical Flags",           sum(r2.flag_count for r2 in results if r2.severity == "critical"), "#,##0"),
            ("Warning Flags",            sum(r2.flag_count for r2 in results if r2.severity == "warning"),  "#,##0"),
            ("Info Flags",               sum(r2.flag_count for r2 in results if r2.severity == "info"),     "#,##0"),
        ]
        for i, (metric, val, fmt) in enumerate(kpis):
            row_fill = FILL_ALT if i % 2 == 1 else None
            _cell(ws, r, 2, metric, bold=True, color=NAVY, fill=row_fill, border=True)
            c = _cell(ws, r, 3, val if not isinstance(val, str) else val,
                      color=DARK, fill=row_fill, border=True)
            if fmt and not isinstance(val, str):
                c.number_format = fmt
            r += 1
        ws.row_dimensions[r].height = 10; r += 1

        # ── Checks Summary ────────────────────────────────────────────────────
        checks_table_start = r
        _section_header(ws, r, 2, "QUALITY CHECKS SUMMARY", span=5); r += 1
        _table_header(ws, r, 2, ["Check", "Issue Type", "Severity", "Flags", "% of Total"])
        r += 1
        chart_data_start = r
        for i, res in enumerate(results):
            sev = res.severity
            pct = f"{res.flag_count / max(total_rows, 1) * 100:.1f}%"
            sev_fill = SEV_FILL.get(sev)
            row_fill = sev_fill if res.flag_count > 0 else (FILL_ALT if i % 2 == 1 else None)
            _cell(ws, r, 2, res.check_name.replace("_", " ").title(), color=DARK, fill=row_fill, border=True)
            _cell(ws, r, 3, res.issue_type, color=DARK, fill=row_fill, border=True)
            _cell(ws, r, 4, sev.upper(), bold=True,
                  color=SEV_COLOR.get(sev, DARK), fill=row_fill, border=True, align="center")
            _cell(ws, r, 5, res.flag_count, color=DARK, fill=row_fill, border=True, align="center")
            _cell(ws, r, 6, pct, color=DARK, fill=row_fill, border=True, align="center")
            r += 1
        chart_data_end = r - 1

        # ── Bar Chart ─────────────────────────────────────────────────────────
        if len(results) > 0 and chart_data_end >= chart_data_start:
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "clustered"
            chart.title = "Flag Count by Check"
            chart.y_axis.title = "Flags"
            chart.x_axis.title = "Check"
            chart.height = 14
            chart.width = 28
            chart.style = 10

            data_ref = Reference(ws,
                                 min_col=5, max_col=5,
                                 min_row=chart_data_start - 1,
                                 max_row=chart_data_end)
            cats_ref = Reference(ws,
                                 min_col=2,
                                 min_row=chart_data_start,
                                 max_row=chart_data_end)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            ws.add_chart(chart, "H5")

    # ── 2. ALL RECORDS ────────────────────────────────────────────────────────

    def _write_all_records(self, wb, results, df):
        if df is None:
            return
        ws = wb.create_sheet("ALL RECORDS")
        ws.sheet_view.showGridLines = False

        sev_map, count_map, checks_map = {}, {}, {}
        for res in results:
            if res.flag_count == 0:
                continue
            for idx in res.flagged_rows.index:
                count_map[idx] = count_map.get(idx, 0) + 1
                cur = sev_map.get(idx, "none")
                if _SEV_RANK[res.severity] > _SEV_RANK[cur]:
                    sev_map[idx] = res.severity
                checks_map.setdefault(idx, []).append(
                    _SHEET.get(res.check_name, res.check_name)
                )

        orig_cols = list(df.columns)
        meta_cols = ["QC Status", "Flag Count", "Max Severity", "Failed Checks"]
        all_cols = orig_cols + meta_cols

        _table_header(ws, 1, 1, all_cols)

        for row_i, (idx, rec) in enumerate(df.iterrows(), start=2):
            flagged = idx in sev_map
            sev = sev_map.get(idx, "none")
            cnt = count_map.get(idx, 0)
            chks = ", ".join(checks_map.get(idx, []))
            status = "FLAGGED" if flagged else "CLEAN"

            if flagged:
                row_fill = SEV_FILL.get(sev)
            else:
                row_fill = FILL_ALT if row_i % 2 == 0 else None

            for ci, col in enumerate(orig_cols, start=1):
                _cell(ws, row_i, ci, rec.get(col), fill=row_fill, border=True)

            base = len(orig_cols) + 1
            status_color = SEV_COLOR.get(sev, GREEN) if flagged else GREEN
            _cell(ws, row_i, base,     status,  bold=flagged, color=status_color,
                  fill=row_fill, border=True, align="center")
            _cell(ws, row_i, base + 1, cnt,     fill=row_fill, border=True, align="center")
            _cell(ws, row_i, base + 2, sev.upper() if flagged else "CLEAN", bold=True,
                  color=status_color,
                  fill=row_fill, border=True, align="center")
            _cell(ws, row_i, base + 3, chks,    fill=row_fill, border=True, wrap=True)

        _auto_width(ws)
        _freeze(ws, row=2)

    # ── 3. FLAGGED RECORDS ────────────────────────────────────────────────────

    def _write_flagged_records(self, wb, results):
        ws = wb.create_sheet("FLAGGED RECORDS")
        ws.sheet_view.showGridLines = False

        all_flagged = []
        for res in results:
            if res.flag_count == 0:
                continue
            df = res.flagged_rows.copy()
            df.insert(0, "_qc_check",   res.check_name)
            df.insert(1, "_severity",   res.severity)
            df.insert(2, "_issue_type", res.issue_type)
            all_flagged.append(df)

        if not all_flagged:
            _cell(ws, 1, 1, "No flagged records found.")
            return

        combined = pd.concat(all_flagged, ignore_index=True)
        cols = list(combined.columns)

        _table_header(ws, 1, 1, cols)

        for row_i, (_, rec) in enumerate(combined.iterrows(), start=2):
            sev = str(rec.get("_severity", "info"))
            fill = SEV_FILL.get(sev) if rec.get("_severity") else (FILL_ALT if row_i % 2 == 0 else None)
            for ci, col in enumerate(cols, start=1):
                _cell(ws, row_i, ci, rec.get(col), fill=fill, border=True)

        _auto_width(ws)
        _freeze(ws, row=2)

    # ── 4. Per-check sheets ───────────────────────────────────────────────────

    def _write_check_sheet(self, wb, result, df_original):
        raw_name = _SHEET.get(result.check_name, result.check_name[:31])
        # Ensure unique sheet name
        existing = wb.sheetnames
        sheet_name = raw_name
        if sheet_name in existing:
            sheet_name = f"{raw_name[:28]}_{sum(1 for s in existing if s.startswith(raw_name[:28]))}"
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False

        sev  = result.severity
        fill = SEV_FILL.get(sev, FILL_BLUE)

        ws.column_dimensions["A"].width = 2
        ws.row_dimensions[1].height = 8
        _section_header(ws, 2, 1, f"  {result.check_name.replace('_', ' ').upper()}", span=8)
        _cell(ws, 3, 1, f"Issue type: {result.issue_type}  |  Severity: {sev.upper()}  |  "
              f"Flags: {result.flag_count}", size=10, color="FF555555")
        ws.row_dimensions[4].height = 8

        flagged = result.flagged_rows

        if flagged.empty:
            _cell(ws, 5, 1, "No records flagged.")
            return

        explanation_cols = [c for c in flagged.columns if c.startswith("_")]
        data_cols = [c for c in flagged.columns if not c.startswith("_")]
        priority_kw = ["id", "interviewer", "instance", "date", "duration", "region", "enumerator"]
        priority_cols = [c for c in data_cols if any(kw in c.lower() for kw in priority_kw)]
        other_cols = [c for c in data_cols if c not in priority_cols]
        display_cols = (priority_cols + other_cols)[:20] + explanation_cols

        if not display_cols:
            display_cols = list(flagged.columns)[:20]

        _table_header(ws, 5, 1, display_cols, fill=FILL_NAVY)

        for row_i, (_, rec) in enumerate(flagged[display_cols].iterrows(), start=6):
            row_fill = fill if row_i % 2 == 0 else None
            for ci, col in enumerate(display_cols, start=1):
                _cell(ws, row_i, ci, rec.get(col), fill=row_fill, border=True)

        _auto_width(ws)
        _freeze(ws, row=6)

    # ── 5. ITVR PERFORMANCE ───────────────────────────────────────────────────

    def _write_itvr_performance(self, wb, results, df):
        int_col = self._detect_col(df, ["interviewer_id", "interviewer", "int_id",
                                         "enumerator_id", "enumerator"])
        if not int_col:
            return

        ws = wb.create_sheet("ITVR PERFORMANCE")
        ws.sheet_view.showGridLines = False

        dur_col  = self._detect_col(df, ["duration_minutes", "duration"])
        date_col = self._detect_col(df, ["interview_date", "date", "survey_date"])

        totals = df.groupby(int_col).size().rename("total_interviews")
        perf = pd.DataFrame(index=totals.index)
        perf.index.name = int_col
        perf = perf.join(totals)

        if dur_col:
            num = pd.to_numeric(df[dur_col], errors="coerce")
            perf = perf.join(
                df.assign(_d=num).groupby(int_col)["_d"]
                .agg(avg_duration="mean", min_duration="min", max_duration="max")
                .round(1)
            )

        if date_col:
            perf = perf.join(
                df.groupby(int_col)[date_col].agg(
                    first_interview="min", last_interview="max"
                )
            )

        for res in results:
            if res.flag_count == 0 or int_col not in res.flagged_rows.columns:
                continue
            short = _SHEET.get(res.check_name, res.check_name)[:20]
            counts = res.flagged_rows[int_col].value_counts().rename(f"flags_{short}")
            perf = perf.join(counts)

        flag_cols = [c for c in perf.columns if c.startswith("flags_")]
        perf["total_flags"] = perf[flag_cols].fillna(0).sum(axis=1).astype(int)
        perf["flag_rate_%"] = (
            perf["total_flags"] / perf["total_interviews"].clip(lower=1) * 100
        ).round(1)

        perf = perf.reset_index().sort_values("flag_rate_%", ascending=False)
        cols = list(perf.columns)

        _table_header(ws, 1, 1, cols)
        for row_i, (_, rec) in enumerate(perf.iterrows(), start=2):
            rate = float(rec.get("flag_rate_%", 0) or 0)
            if rate >= 30:
                row_fill = FILL_RED
            elif rate >= 10:
                row_fill = FILL_ORANGE
            else:
                row_fill = FILL_ALT if row_i % 2 == 0 else None
            for ci, col in enumerate(cols, start=1):
                _cell(ws, row_i, ci, rec.get(col), fill=row_fill, border=True)

        _auto_width(ws)
        _freeze(ws, row=2)

    # ── 6. PRODUCTIVITY ───────────────────────────────────────────────────────

    def _write_productivity(self, wb, df):
        int_col  = self._detect_col(df, ["interviewer_id", "interviewer", "int_id",
                                          "enumerator_id", "enumerator"])
        date_col = self._detect_col(df, ["interview_date", "date", "survey_date"])
        if not int_col or not date_col:
            return

        ws = wb.create_sheet("PRODUCTIVITY")
        ws.sheet_view.showGridLines = False

        work = df[[int_col, date_col]].copy()
        work[date_col] = work[date_col].astype(str).str[:10]
        pivot = (
            work.groupby([date_col, int_col])
            .size()
            .unstack(fill_value=0)
            .reset_index()
        )

        interviewers = [c for c in pivot.columns if c != date_col]
        headers = [date_col] + interviewers + ["Daily Total"]
        _table_header(ws, 1, 1, headers)

        last_row = 1
        for row_i, (_, rec) in enumerate(pivot.iterrows(), start=2):
            row_total = int(sum(rec.get(i, 0) for i in interviewers))
            alt_fill = FILL_ALT if row_i % 2 == 0 else None
            _cell(ws, row_i, 1, str(rec[date_col]), fill=alt_fill, border=True)
            for ci, itvr in enumerate(interviewers, start=2):
                val = int(rec.get(itvr, 0))
                cell_fill = FILL_GREEN if val >= 3 else (FILL_ORANGE if val == 1 else alt_fill)
                _cell(ws, row_i, ci, val, fill=cell_fill, border=True, align="center")
            _cell(ws, row_i, len(interviewers) + 2, row_total, bold=True,
                  fill=alt_fill, border=True, align="center")
            last_row = row_i

        # Column totals row
        total_row = last_row + 2
        _cell(ws, total_row, 1, "TOTAL", bold=True, color=WHITE, fill=FILL_NAVY, border=True)
        col_totals = [int(pivot[i].sum()) for i in interviewers]
        for ci, tot in enumerate(col_totals, start=2):
            _cell(ws, total_row, ci, tot, bold=True, color=WHITE,
                  fill=FILL_NAVY, border=True, align="center")
        _cell(ws, total_row, len(interviewers) + 2, sum(col_totals),
              bold=True, color=WHITE, fill=FILL_NAVY, border=True, align="center")

        _auto_width(ws, max_w=20)
        _freeze(ws, row=2)

    # ── 7. RISK SCORECARD ─────────────────────────────────────────────────────

    def _write_risk_scorecard(self, wb, results, df):
        int_col = self._detect_col(df, ["interviewer_id", "interviewer", "int_id",
                                         "enumerator_id", "enumerator"])
        if not int_col:
            return

        dur_col = self._detect_col(df, ["duration_minutes", "duration"])

        totals = df.groupby(int_col).size().rename("total_interviews")
        perf = pd.DataFrame(index=totals.index)
        perf.index.name = int_col
        perf = perf.join(totals)

        for key in _RISK_WEIGHTS:
            perf[f"{key}_flags"] = 0

        for res in results:
            category = _RISK_FLAG_MAP.get(res.check_name)
            if not category or res.flag_count == 0 or int_col not in res.flagged_rows.columns:
                continue
            counts = res.flagged_rows[int_col].value_counts()
            for idx in perf.index:
                perf.loc[idx, f"{category}_flags"] += int(counts.get(idx, 0))

        for key in _RISK_WEIGHTS:
            perf[f"{key}_rate"] = (
                perf[f"{key}_flags"] / perf["total_interviews"].clip(lower=1)
            ).clip(0, 1)

        perf["risk_score"] = (
            perf["fabrication_rate"]    * _RISK_WEIGHTS["fabrication"]    +
            perf["duration_rate"]       * _RISK_WEIGHTS["duration"]       +
            perf["straightlining_rate"] * _RISK_WEIGHTS["straightlining"] +
            perf["productivity_rate"]   * _RISK_WEIGHTS["productivity"]
        ).mul(100).round(1)

        perf["total_flags"] = (
            perf["fabrication_flags"] + perf["duration_flags"] +
            perf["straightlining_flags"] + perf["productivity_flags"]
        ).astype(int)
        perf["flag_rate_%"] = (
            perf["total_flags"] / perf["total_interviews"].clip(lower=1) * 100
        ).round(1)

        if dur_col:
            num = pd.to_numeric(df[dur_col], errors="coerce")
            perf = perf.join(
                df.assign(_d=num).groupby(int_col)["_d"]
                .agg(avg_duration="mean").round(1)
            )

        perf = perf.reset_index().sort_values("risk_score", ascending=False)

        ws = wb.create_sheet("RISK SCORECARD")
        ws.sheet_view.showGridLines = False

        ws.column_dimensions["A"].width = 2
        ws.row_dimensions[1].height = 8
        _section_header(ws, 2, 1, "  INTERVIEWER RISK SCORECARD", span=12)
        _cell(ws, 3, 1,
              "RED ≥ 30 risk score  |  AMBER ≥ 15  |  GREEN < 15  "
              "(weights: fabrication 40 %, duration 25 %, straightlining 25 %, productivity 10 %)",
              size=9, color="FF555555")
        ws.row_dimensions[4].height = 8

        display_cols = [
            int_col, "total_interviews",
            "fabrication_flags", "duration_flags",
            "straightlining_flags", "productivity_flags",
            "total_flags", "flag_rate_%", "risk_score",
        ]
        if dur_col and "avg_duration" in perf.columns:
            display_cols.append("avg_duration")

        headers = [c.replace("_", " ").title() for c in display_cols]
        _table_header(ws, 5, 1, headers)

        for row_i, (_, rec) in enumerate(perf.iterrows(), start=6):
            score = float(rec.get("risk_score", 0) or 0)
            if score >= 30:
                row_fill = FILL_RED
            elif score >= 15:
                row_fill = FILL_ORANGE
            else:
                row_fill = FILL_GREEN if row_i % 2 == 0 else None

            for ci, col in enumerate(display_cols, start=1):
                _cell(ws, row_i, ci, rec.get(col),
                      fill=row_fill, border=True,
                      align="center" if ci > 1 else "left")

        _auto_width(ws)
        _freeze(ws, row=6)

    # ── Utilities ─────────────────────────────────────────────────────────────

    @staticmethod
    def _count_unique_flagged_rows(results):
        flagged = set()
        for res in results:
            if res.flag_count > 0:
                flagged.update(res.flagged_rows.index.tolist())
        return len(flagged)

    @staticmethod
    def _detect_col(df, candidates):
        if df is None:
            return None
        lower = {c.lower(): c for c in df.columns}
        for candidate in candidates:
            if candidate.lower() in lower:
                return lower[candidate.lower()]
        return None

    def print_summary(self, results: List[CheckResult]):
        print("\n" + "=" * 60)
        print("QC ENGINE RESULTS SUMMARY")
        print("=" * 60)
        total = 0
        for r in results:
            icon = "[CRIT]" if r.severity == "critical" else "[WARN]" if r.severity == "warning" else "[INFO]"
            print(f"{icon} {r.check_name}: {r.flag_count} flags")
            total += r.flag_count
        print("-" * 60)
        print(f"TOTAL FLAGS: {total}")
        print("=" * 60 + "\n")


# ── Standalone helper: append one supplemental sheet to an existing workbook ──

def append_supplemental_sheet(report_path: str, entry: dict) -> None:
    """
    Open an existing QC workbook and add one supplemental check as a new sheet.
    Safe to call multiple times — deduplicates sheet names.

    entry keys: check_name, issue_type, severity, flag_count, flagged_rows (list of dicts)
    """
    wb = load_workbook(report_path)

    check_name = entry.get("check_name", "custom_check")
    issue_type = entry.get("issue_type", "Custom Check")
    sev        = entry.get("severity", "warning")
    flag_count = entry.get("flag_count", 0)
    rows       = entry.get("flagged_rows", [])

    # Build a unique sheet name
    raw = _SHEET.get(check_name, check_name.upper().replace("_", " "))[:28]
    sheet_name = raw
    existing = wb.sheetnames
    if sheet_name in existing:
        idx = sum(1 for s in existing if s.startswith(raw[:24]))
        sheet_name = f"{raw[:24]}_{idx}"

    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    fill = SEV_FILL.get(sev, FILL_BLUE)

    ws.column_dimensions["A"].width = 2
    ws.row_dimensions[1].height = 8
    _section_header(ws, 2, 1, f"  {check_name.replace('_', ' ').upper()}", span=8)
    _cell(ws, 3, 1,
          f"Issue type: {issue_type}  |  Severity: {sev.upper()}  |  Flags: {flag_count}",
          size=10, color="FF555555")
    ws.row_dimensions[4].height = 8

    if not rows:
        _cell(ws, 5, 1, "No records flagged.")
        wb.save(report_path)
        return

    display_cols = [k for k in rows[0].keys() if not k.startswith("_")]
    internal_cols = [k for k in rows[0].keys() if k.startswith("_")]
    all_cols = display_cols + internal_cols
    if not all_cols:
        all_cols = list(rows[0].keys())[:30]

    _table_header(ws, 5, 1, all_cols, fill=FILL_NAVY)

    for row_i, rec in enumerate(rows, start=6):
        row_fill = fill if row_i % 2 == 0 else None
        for ci, col in enumerate(all_cols, start=1):
            _cell(ws, row_i, ci, rec.get(col), fill=row_fill, border=True)

    _auto_width(ws)
    _freeze(ws, row=6)

    wb.save(report_path)
    logger.info(f"Supplemental sheet '{sheet_name}' appended to {report_path}")
