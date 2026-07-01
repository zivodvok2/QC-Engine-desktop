"""
Dashboard API router — serves the React dashboard UI.
Mirrors the Streamlit dashboard/pages_modules logic via the shared SQLite DB.
"""
import io
import random
import uuid
from datetime import date, timedelta
from typing import Any, Dict, List, Optional

import bcrypt
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from pydantic import BaseModel

import shared_db
from routers.auth import get_current_user

router = APIRouter(prefix="/dashboard")

# ── Role constants ─────────────────────────────────────────────────────────────

ADMIN_ROLES = {"qc_executive", "operations_manager"}
UPLOAD_ROLES = {"qc_executive", "operations_manager", "qc_officer"}

# ── Backcheck column map (from dashboard/config.py) ────────────────────────────

BACKCHECK_IFIELD_COL_MAP = {
    "BC instance ID": "bc_instance_id",
    "Interview Status": "interview_status",
    "Original interview instance ID": "original_instance_id",
    "Region": "region",
    "Location": "location",
    "Sample Point ID": "sample_point_id",
    "Back-checker ID": "backchecker_id",
    "Interviewer ID": "interviewer_id",
    "Script Name": "script_name",
    "Interview Date": "interview_date",
    "Interview Start Time": "interview_start_time",
    "Back-check completion date": "backcheck_date",
    "Back-check completion time": "backcheck_time",
    "(1) Totally fraudulent intervi": "error_01",
    "(2) Different respondent name ": "error_02",
    "(3) Mismatched quota(s) e.g. d": "error_03",
    "(4) Wrong usership quotas and/": "error_04",
    "(5) Wrong telephone number": "error_05",
    "(6) Unattainable/Invalid telep": "error_06",
    "(7) Engaged/Answer machine": "error_07",
    "(8) Refused/Unable to speak to": "error_08",
    "(9) Respondent does not rememb": "error_09",
    "(10) Back-check abandoned/inco": "error_10",
    "(11) CAPI interview done on Pa": "error_11",
    "(12) Voice recording permission": "error_12",
    "(13) Respondent has already pa": "error_13",
}

# ── Auth helpers ───────────────────────────────────────────────────────────────

def _require_admin(user: dict):
    if user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin role required")


def _require_upload(user: dict):
    if user["role"] not in UPLOAD_ROLES:
        raise HTTPException(status_code=403, detail="Upload role required")


def _require_db():
    if not shared_db.db_available():
        raise HTTPException(status_code=503, detail="Database not available")


# ── Normalise backcheck Excel ──────────────────────────────────────────────────

def _normalise_backcheck(df: pd.DataFrame) -> list:
    """Apply BACKCHECK_IFIELD_COL_MAP, parse dates, coerce error cols to int."""
    df = df.rename(columns=BACKCHECK_IFIELD_COL_MAP)
    # Lowercase remaining columns
    df.columns = [str(c).strip() for c in df.columns]

    # Parse date columns
    for col in ("interview_date", "backcheck_date"):
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime("%Y-%m-%d")

    # Coerce error columns
    for i in range(1, 14):
        col = f"error_{i:02d}"
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

    records = []
    for _, row in df.iterrows():
        rec = {}
        for col in df.columns:
            v = row[col]
            if hasattr(v, "item"):
                v = v.item()
            rec[col] = None if (v != v) else v  # nan → None
        records.append(rec)
    return records


def _normalise_listen_in(df: pd.DataFrame) -> list:
    """Normalise listen-in batch upload."""
    df.columns = [str(c).strip().lower() for c in df.columns]
    rename = {
        "instance id": "instance_id",
        "interviewer id": "interviewer_id",
        "listen date": "listen_date",
        "listen type": "listen_type",
        "issues noted": "issues_noted",
        "action taken": "action_taken",
    }
    df = df.rename(columns=rename)
    if "listen_date" in df.columns:
        df["listen_date"] = pd.to_datetime(df["listen_date"], errors="coerce").dt.strftime("%Y-%m-%d")

    records = []
    for _, row in df.iterrows():
        rec = {}
        for col in df.columns:
            v = row[col]
            if hasattr(v, "item"):
                v = v.item()
            rec[col] = None if (v != v) else v
        records.append(rec)
    return records


# ── Pydantic schemas ───────────────────────────────────────────────────────────

class CreateUserBody(BaseModel):
    email: str
    password: str
    full_name: str
    role: str


class UpdateRoleBody(BaseModel):
    role: str


class ToggleActiveBody(BaseModel):
    is_active: bool


class CreateProjectFullBody(BaseModel):
    name: str
    client: Optional[str] = None
    job_number: Optional[str] = None
    sample_target: int = 0
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    backcheck_target: float = 0.20
    listenin_target: float = 0.10
    accompaniment_target: float = 0.20
    loi_min_minutes: float = 0
    loi_pct_threshold: float = 0.50
    flag_warning_pct: float = 5.0
    flag_critical_pct: float = 10.0


class UpdateProjectBody(BaseModel):
    name: Optional[str] = None
    client: Optional[str] = None
    job_number: Optional[str] = None
    sample_target: Optional[int] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    status: Optional[str] = None
    backcheck_target: Optional[float] = None
    listenin_target: Optional[float] = None
    accompaniment_target: Optional[float] = None
    loi_min_minutes: Optional[float] = None
    loi_pct_threshold: Optional[float] = None
    flag_warning_pct: Optional[float] = None
    flag_critical_pct: Optional[float] = None


class AssignUserBody(BaseModel):
    user_id: int


class CreateSupervisorBody(BaseModel):
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    region: Optional[str] = None


class UpsertInterviewerBody(BaseModel):
    interviewer_code: str
    name: Optional[str] = None
    supervisor_id: Optional[int] = None
    region: Optional[str] = None
    is_active: int = 1


class ManualListenInBody(BaseModel):
    instance_id: Optional[str] = None
    interviewer_id: str
    region: Optional[str] = None
    listen_date: str
    listen_type: str
    result: str
    issues_noted: Optional[str] = None
    action_taken: Optional[str] = None


# ── Summary & Activity ─────────────────────────────────────────────────────────

@router.get("/summary")
def get_summary(user: dict = Depends(get_current_user)):
    _require_db()
    return shared_db.get_dashboard_summary()


@router.get("/activity")
def get_activity(
    limit: int = 20,
    user: dict = Depends(get_current_user),
):
    _require_db()
    return shared_db.get_project_activity(limit=limit)


# ── Per-project data endpoints ─────────────────────────────────────────────────

@router.get("/projects/{project_id}/quality-records")
def get_quality_records(project_id: int, user: dict = Depends(get_current_user)):
    _require_db()
    return shared_db.get_quality_records(project_id)


@router.get("/projects/{project_id}/backcheck-records")
def get_backcheck_records(project_id: int, user: dict = Depends(get_current_user)):
    _require_db()
    return shared_db.get_backcheck_records(project_id)


@router.get("/projects/{project_id}/listen-in-records")
def get_listen_in_records(project_id: int, user: dict = Depends(get_current_user)):
    _require_db()
    return shared_db.get_listen_in_records(project_id)


@router.get("/projects/{project_id}/performance-records")
def get_performance_records(project_id: int, user: dict = Depends(get_current_user)):
    _require_db()
    return shared_db.get_performance_records(project_id)


@router.get("/projects/{project_id}/timing-records")
def get_timing_records(project_id: int, user: dict = Depends(get_current_user)):
    _require_db()
    return shared_db.get_timing_records(project_id)


@router.get("/projects/{project_id}/cancelled-records")
def get_cancelled_records(project_id: int, user: dict = Depends(get_current_user)):
    _require_db()
    return shared_db.get_cancelled_records(project_id)


@router.get("/projects/{project_id}/upload-log")
def get_upload_log(project_id: int, user: dict = Depends(get_current_user)):
    _require_db()
    return shared_db.get_upload_log(project_id)


# ── Backcheck file upload ──────────────────────────────────────────────────────

@router.post("/projects/{project_id}/backcheck")
async def upload_backcheck(
    project_id: int,
    file: UploadFile = File(...),
    wave_label: Optional[str] = Form(None),
    user: dict = Depends(get_current_user),
):
    _require_db()
    _require_upload(user)
    project = shared_db.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    content = await file.read()
    try:
        df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {e}")

    records = _normalise_backcheck(df)
    uid = shared_db.insert_backcheck_records(
        project_id, user["id"], file.filename or "upload.xlsx", records, wave_label
    )
    return {"status": "ok", "upload_id": uid, "row_count": len(records)}


# ── Listen-in endpoints ────────────────────────────────────────────────────────

@router.post("/projects/{project_id}/listen-in")
async def upload_listen_in(
    project_id: int,
    file: UploadFile = File(...),
    wave_label: Optional[str] = Form(None),
    user: dict = Depends(get_current_user),
):
    _require_db()
    _require_upload(user)
    project = shared_db.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    content = await file.read()
    try:
        df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {e}")

    records = _normalise_listen_in(df)
    uid = shared_db.insert_listen_in_batch(
        project_id, user["id"], file.filename or "upload.xlsx", records, wave_label
    )
    return {"status": "ok", "upload_id": uid, "row_count": len(records)}


@router.post("/projects/{project_id}/listen-in/manual")
def add_manual_listen_in(
    project_id: int,
    body: ManualListenInBody,
    user: dict = Depends(get_current_user),
):
    _require_db()
    _require_upload(user)
    project = shared_db.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    shared_db.insert_listen_in_record(
        project_id=project_id,
        logged_by=user["id"],
        instance_id=body.instance_id,
        interviewer_id=body.interviewer_id,
        region=body.region,
        listen_date=body.listen_date,
        listen_type=body.listen_type,
        result=body.result,
        issues_noted=body.issues_noted,
        action_taken=body.action_taken,
    )
    return {"status": "ok"}


@router.post("/projects/{project_id}/performance")
async def upload_performance(
    project_id: int,
    file: UploadFile = File(...),
    wave_label: Optional[str] = Form(None),
    user: dict = Depends(get_current_user),
):
    _require_db()
    _require_upload(user)
    project = shared_db.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    content = await file.read()
    try:
        df = pd.read_excel(io.BytesIO(content), skiprows=2)
    except Exception:
        try:
            df = pd.read_excel(io.BytesIO(content))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {e}")
    from dashboard.config import PERFORMANCE_IFIELD_COL_MAP
    df = df.rename(columns=PERFORMANCE_IFIELD_COL_MAP)
    df.columns = [str(c).strip().lower() for c in df.columns]
    for col in ("first_interview", "last_interview"):
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime("%Y-%m-%d")
    for col in ["interview_completes", "followup_completes", "ecs_completes", "work_summary",
                "accompaniments", "cancelled_interviews", "backcheck_telephone_created",
                "backcheck_f2f_created", "backcheck_f2f_infield", "backcheck_total", "backcheck_completed"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
    records = [{c: (None if (v != v) else (v.item() if hasattr(v, "item") else v))
                for c, v in row.items()} for _, row in df.iterrows()]
    uid = shared_db.insert_performance_records(
        project_id, user["id"], file.filename or "upload.xlsx", records, wave_label
    )
    return {"status": "ok", "upload_id": uid, "row_count": len(records)}


@router.post("/projects/{project_id}/timing")
async def upload_timing(
    project_id: int,
    file: UploadFile = File(...),
    wave_label: Optional[str] = Form(None),
    user: dict = Depends(get_current_user),
):
    _require_db()
    _require_upload(user)
    project = shared_db.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    content = await file.read()
    try:
        df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {e}")
    df.columns = [str(c).strip().lower() for c in df.columns]
    rename = {
        "instance id": "instance_id", "interviewer id": "interviewer_id",
        "interview date": "interview_date", "duration minutes": "duration_minutes",
        "duration_mins": "duration_minutes", "loi": "duration_minutes",
    }
    df = df.rename(columns=rename)
    if "interview_date" in df.columns:
        df["interview_date"] = pd.to_datetime(df["interview_date"], errors="coerce").dt.strftime("%Y-%m-%d")
    if "duration_minutes" in df.columns:
        df["duration_minutes"] = pd.to_numeric(df["duration_minutes"], errors="coerce")
    records = [{c: (None if (v != v) else (v.item() if hasattr(v, "item") else v))
                for c, v in row.items()} for _, row in df.iterrows()]
    uid = shared_db.insert_timing_records(
        project_id, user["id"], file.filename or "upload.xlsx", records, wave_label
    )
    return {"status": "ok", "upload_id": uid, "row_count": len(records)}


@router.post("/projects/{project_id}/cancelled")
async def upload_cancelled(
    project_id: int,
    file: UploadFile = File(...),
    wave_label: Optional[str] = Form(None),
    user: dict = Depends(get_current_user),
):
    _require_db()
    _require_upload(user)
    project = shared_db.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    content = await file.read()
    try:
        df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {e}")
    from dashboard.config import CANCELLED_IFIELD_COL_MAP
    df = df.rename(columns=CANCELLED_IFIELD_COL_MAP)
    df.columns = [str(c).strip().lower() for c in df.columns]
    if "interview_date" in df.columns:
        df["interview_date"] = pd.to_datetime(df["interview_date"], errors="coerce").dt.strftime("%Y-%m-%d")
    for col in ["interview_length", "active_length", "avg_length_sample_point",
                "avg_length_region", "avg_length_project", "idle_time", "gap_to_last"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    records = [{c: (None if (v != v) else (v.item() if hasattr(v, "item") else v))
                for c, v in row.items()} for _, row in df.iterrows()]
    uid = shared_db.insert_cancelled_records(
        project_id, user["id"], file.filename or "upload.xlsx", records, wave_label
    )
    return {"status": "ok", "upload_id": uid, "row_count": len(records)}


@router.delete("/projects/{project_id}/listen-in/{record_id}")
def delete_listen_in(
    project_id: int,
    record_id: int,
    user: dict = Depends(get_current_user),
):
    _require_db()
    _require_upload(user)
    shared_db.delete_listen_in_record(record_id)
    return {"status": "ok"}


# ── Upload management ──────────────────────────────────────────────────────────

@router.get("/uploads")
def list_all_uploads(user: dict = Depends(get_current_user)):
    _require_db()
    _require_admin(user)
    return shared_db.get_upload_log(None)


@router.delete("/uploads/{upload_id}")
def delete_upload(
    upload_id: str,
    report_type: str,
    user: dict = Depends(get_current_user),
):
    _require_db()
    _require_admin(user)
    if shared_db.is_upload_locked(upload_id):
        raise HTTPException(status_code=403, detail="This upload is locked and cannot be deleted")
    shared_db.delete_upload(upload_id, report_type)
    return {"status": "ok"}


# ── Users (admin only) ─────────────────────────────────────────────────────────

@router.get("/users")
def list_users(user: dict = Depends(get_current_user)):
    _require_db()
    _require_admin(user)
    return shared_db.get_all_users()


@router.post("/users")
def create_user(body: CreateUserBody, user: dict = Depends(get_current_user)):
    _require_db()
    _require_admin(user)
    pw_hash = bcrypt.hashpw(body.password.encode(), bcrypt.gensalt()).decode()
    ok, err = shared_db.create_user(
        email=body.email.strip().lower(),
        password_hash=pw_hash,
        full_name=body.full_name,
        role=body.role,
        created_by=user["id"],
    )
    if not ok:
        raise HTTPException(status_code=400, detail=err)
    return {"status": "ok"}


@router.put("/users/{user_id}/role")
def update_role(user_id: int, body: UpdateRoleBody, user: dict = Depends(get_current_user)):
    _require_db()
    _require_admin(user)
    shared_db.update_user_role(user_id, body.role)
    return {"status": "ok"}


@router.patch("/users/{user_id}/active")
def toggle_active(user_id: int, body: ToggleActiveBody, user: dict = Depends(get_current_user)):
    _require_db()
    _require_admin(user)
    shared_db.toggle_user_active(user_id, 1 if body.is_active else 0)
    return {"status": "ok"}


# ── Projects (admin only for create/update) ────────────────────────────────────

@router.post("/projects/full")
def create_project_full(body: CreateProjectFullBody, user: dict = Depends(get_current_user)):
    _require_db()
    _require_admin(user)
    existing = shared_db.get_project_by_name(body.name.strip())
    if existing:
        raise HTTPException(status_code=409, detail=f"A project named '{body.name.strip()}' already exists")
    project_id = shared_db.create_project_full(
        name=body.name,
        client=body.client,
        sample_target=body.sample_target,
        start_date=body.start_date,
        end_date=body.end_date,
        backcheck_target=body.backcheck_target,
        listenin_target=body.listenin_target,
        accompaniment_target=body.accompaniment_target,
        created_by=user["id"],
        job_number=body.job_number,
        loi_min_minutes=body.loi_min_minutes,
        loi_pct_threshold=body.loi_pct_threshold,
        flag_warning_pct=body.flag_warning_pct,
        flag_critical_pct=body.flag_critical_pct,
    )
    return shared_db.get_project(project_id)


@router.put("/projects/{project_id}")
def update_project(
    project_id: int,
    body: UpdateProjectBody,
    user: dict = Depends(get_current_user),
):
    _require_db()
    _require_admin(user)
    project = shared_db.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    kwargs = {k: v for k, v in body.model_dump().items() if v is not None}
    shared_db.update_project(project_id, **kwargs)
    return shared_db.get_project(project_id)


# ── Project assignments ────────────────────────────────────────────────────────

@router.get("/projects/{project_id}/assignees")
def get_assignees(project_id: int, user: dict = Depends(get_current_user)):
    _require_db()
    return shared_db.get_project_assignees(project_id)


@router.post("/projects/{project_id}/assignees")
def add_assignee(
    project_id: int,
    body: AssignUserBody,
    user: dict = Depends(get_current_user),
):
    _require_db()
    _require_admin(user)
    shared_db.assign_user_to_project(project_id, body.user_id, user["id"])
    return {"status": "ok"}


@router.delete("/projects/{project_id}/assignees/{target_user_id}")
def remove_assignee(
    project_id: int,
    target_user_id: int,
    user: dict = Depends(get_current_user),
):
    _require_db()
    _require_admin(user)
    shared_db.remove_user_from_project(project_id, target_user_id)
    return {"status": "ok"}


# ── Download templates ─────────────────────────────────────────────────────────

@router.get("/templates/{template_type}")
def download_template(
    template_type: str,
    user: dict = Depends(get_current_user),
):
    from fastapi.responses import StreamingResponse

    TEMPLATES = {
        "backcheck": {
            "filename": "backcheck_template.xlsx",
            "columns": [
                "BC_INSTANCE_ID", "ORIGINAL_INSTANCE_ID", "INTERVIEW_STATUS",
                "REGION", "LOCATION", "SAMPLE_POINT_ID",
                "BACKCHECKER_ID", "INTERVIEWER_ID", "SCRIPT_NAME",
                "INTERVIEW_DATE", "INTERVIEW_START_TIME", "BACKCHECK_DATE", "BACKCHECK_TIME",
            ] + [f"ERROR_{i:02d}" for i in range(1, 14)],
            "example": {
                "BC_INSTANCE_ID": "3782466", "ORIGINAL_INSTANCE_ID": "3782413",
                "INTERVIEW_STATUS": "Completed - Pending Export",
                "REGION": "Nairobi", "LOCATION": "Githunguri road", "SAMPLE_POINT_ID": "NB09",
                "BACKCHECKER_ID": "LE0018", "INTERVIEWER_ID": "NB0664",
                "SCRIPT_NAME": "Project Name", "INTERVIEW_DATE": "2025-04-15",
                "INTERVIEW_START_TIME": "14:56:22", "BACKCHECK_DATE": "2025-04-16",
                "BACKCHECK_TIME": "08:50",
                **{f"ERROR_{i:02d}": 0 for i in range(1, 14)},
            },
        },
        "listen_in": {
            "filename": "listen_in_template.xlsx",
            "columns": ["INSTANCE_ID", "INTERVIEWER_ID", "REGION", "LISTEN_DATE", "LISTEN_TYPE", "RESULT", "ISSUES_NOTED", "ACTION_TAKEN"],
            "example": {
                "INSTANCE_ID": "3782322", "INTERVIEWER_ID": "NB0664",
                "REGION": "Nairobi", "LISTEN_DATE": "2025-04-15",
                "LISTEN_TYPE": "Telephone", "RESULT": "Pass",
                "ISSUES_NOTED": "", "ACTION_TAKEN": "",
            },
        },
        "quality_report": {
            "filename": "quality_report_template.xlsx",
            "columns": [
                "INSTANCE_ID", "INTERVIEWER_ID", "INTERVIEW_DATE", "START_TIME", "END_TIME",
                "DURATION_MINUTES", "DURATION_VALIDATION", "DURATION_FLAG", "STRAIGHT_LINING",
                "LONG_PAUSE", "GPS_STATUS", "PHONE_PRESENT", "AUDIO_PRESENT",
                "REGION", "LOCATION", "SAMPLE_POINT_ID", "APPROVAL_STATUS", "QC_COMMENTS",
            ],
            "example": {
                "INSTANCE_ID": "3782413", "INTERVIEWER_ID": "NB0664",
                "INTERVIEW_DATE": "2025-04-15", "START_TIME": "14:56:22", "END_TIME": "15:38:05",
                "DURATION_MINUTES": 41.7, "DURATION_VALIDATION": "Pass", "DURATION_FLAG": "Pass",
                "STRAIGHT_LINING": "Pass", "LONG_PAUSE": "Pass", "GPS_STATUS": "Pass",
                "PHONE_PRESENT": "Yes", "AUDIO_PRESENT": "Yes",
                "REGION": "Nairobi", "LOCATION": "Githunguri road", "SAMPLE_POINT_ID": "NB09",
                "APPROVAL_STATUS": "Approved", "QC_COMMENTS": "",
            },
        },
    }

    TEMPLATES.update({
        "performance": {
            "filename": "performance_template.xlsx",
            "columns": ["Login", "Management Region", "Interview completes", "Follow-Ups completes",
                        "ECS completes", "Work Summary", "Accompaniment", "Cancelled Interviews",
                        "Back-Checks\ncompleted"],
            "example": {"Login": "NB0664", "Management Region": "Nairobi",
                        "Interview completes": 12, "Follow-Ups completes": 0,
                        "ECS completes": 0, "Work Summary": 12, "Accompaniment": 2,
                        "Cancelled Interviews": 1, "Back-Checks\ncompleted": 3},
        },
        "timing": {
            "filename": "timing_template.xlsx",
            "columns": ["INSTANCE_ID", "INTERVIEWER_ID", "REGION", "INTERVIEW_DATE", "DURATION_MINUTES"],
            "example": {"INSTANCE_ID": "3782413", "INTERVIEWER_ID": "NB0664",
                        "REGION": "Nairobi", "INTERVIEW_DATE": "2025-04-15", "DURATION_MINUTES": 41.7},
        },
        "cancelled": {
            "filename": "cancelled_template.xlsx",
            "columns": ["Instance ID ", "Interviewer ID", "Region", "Interview Date",
                        "Interview Start Time", "Interview End Time", "Interview length"],
            "example": {"Instance ID ": "3782413", "Interviewer ID": "NB0664",
                        "Region": "Nairobi", "Interview Date": "2025-04-15",
                        "Interview Start Time": "14:56:22", "Interview End Time": "15:38:05",
                        "Interview length": 41.7},
        },
    })

    if template_type not in TEMPLATES:
        raise HTTPException(status_code=404, detail=f"Unknown template type: {template_type}")

    tpl = TEMPLATES[template_type]
    df = pd.DataFrame(columns=tpl["columns"])
    df.loc[0] = {c: tpl["example"].get(c, "") for c in tpl["columns"]}

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Template")
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{tpl["filename"]}"'},
    )


# ── Combined project report (multi-sheet Excel) ────────────────────────────────

@router.get("/projects/{project_id}/combined-report")
def download_combined_report(project_id: int, user: dict = Depends(get_current_user)):
    from fastapi.responses import StreamingResponse

    _require_db()
    project = shared_db.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    quality = shared_db.get_quality_records(project_id)
    backcheck = shared_db.get_backcheck_records(project_id)
    listen_in = shared_db.get_listen_in_records(project_id)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        # QC Summary sheet
        if quality:
            df_q = pd.DataFrame(quality)
            # Drop internal ID columns for the report
            drop_cols = [c for c in ("id", "project_id", "upload_id", "uploaded_by", "extra_data") if c in df_q.columns]
            df_q = df_q.drop(columns=drop_cols)
            df_q.to_excel(w, index=False, sheet_name="Quality Report")

            # Flag summary
            flag_summary = pd.DataFrame([{
                "Metric": "Total Records", "Value": len(df_q),
            }, {
                "Metric": "Approved",
                "Value": (df_q["approval_status"] == "Approved").sum() if "approval_status" in df_q.columns else 0,
            }, {
                "Metric": "Cancelled",
                "Value": (df_q["approval_status"] == "Cancelled").sum() if "approval_status" in df_q.columns else 0,
            }, {
                "Metric": "Duration Flags",
                "Value": (df_q["duration_flag"] == "Flag").sum() if "duration_flag" in df_q.columns else 0,
            }, {
                "Metric": "Straight-lining Flags",
                "Value": (df_q["straight_lining"] == "Flag").sum() if "straight_lining" in df_q.columns else 0,
            }])
            flag_summary.to_excel(w, index=False, sheet_name="QC Summary")
        else:
            pd.DataFrame({"Note": ["No quality records yet"]}).to_excel(w, index=False, sheet_name="Quality Report")

        # Backcheck sheet
        if backcheck:
            df_b = pd.DataFrame(backcheck)
            drop_cols = [c for c in ("id", "project_id", "upload_id", "uploaded_by") if c in df_b.columns]
            df_b = df_b.drop(columns=drop_cols)
            df_b.to_excel(w, index=False, sheet_name="Back-check")
        else:
            pd.DataFrame({"Note": ["No back-check records yet"]}).to_excel(w, index=False, sheet_name="Back-check")

        # Listen-in sheet
        if listen_in:
            df_l = pd.DataFrame(listen_in)
            drop_cols = [c for c in ("id", "project_id", "upload_id", "logged_by") if c in df_l.columns]
            df_l = df_l.drop(columns=drop_cols)
            df_l.to_excel(w, index=False, sheet_name="Listen-in")
        else:
            pd.DataFrame({"Note": ["No listen-in records yet"]}).to_excel(w, index=False, sheet_name="Listen-in")

        # Interviewer cross-tab (from quality records)
        if quality:
            df_q2 = pd.DataFrame(quality)
            if "interviewer_id" in df_q2.columns:
                itv_summary = df_q2.groupby("interviewer_id").agg(
                    interviews=("id", "count"),
                    dur_flags=("duration_flag", lambda x: (x == "Flag").sum()),
                    sl_flags=("straight_lining", lambda x: (x == "Flag").sum()),
                    approved=("approval_status", lambda x: (x == "Approved").sum()),
                ).reset_index()
                itv_summary.columns = ["Interviewer ID", "Interviews", "Duration Flags", "SL Flags", "Approved"]
                itv_summary.to_excel(w, index=False, sheet_name="Interviewers")

    pname = (project.get("name") or "project").replace(" ", "_")[:30]
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{pname}_combined_report.xlsx"'},
    )


# ── Supervisors ────────────────────────────────────────────────────────────────

@router.get("/supervisors")
def list_supervisors(user: dict = Depends(get_current_user)):
    _require_db()
    return shared_db.get_all_supervisors()


@router.post("/supervisors")
def create_supervisor(body: CreateSupervisorBody, user: dict = Depends(get_current_user)):
    _require_db()
    _require_admin(user)
    sid = shared_db.create_supervisor(body.name, body.email, body.phone, body.region)
    return {"id": sid, **body.model_dump()}


@router.put("/supervisors/{supervisor_id}")
def update_supervisor(supervisor_id: int, body: CreateSupervisorBody, user: dict = Depends(get_current_user)):
    _require_db()
    _require_admin(user)
    shared_db.update_supervisor(supervisor_id, **{k: v for k, v in body.model_dump().items() if v is not None})
    return {"status": "ok"}


@router.delete("/supervisors/{supervisor_id}")
def delete_supervisor(supervisor_id: int, user: dict = Depends(get_current_user)):
    _require_db()
    _require_admin(user)
    shared_db.delete_supervisor(supervisor_id)
    return {"status": "ok"}


# ── Interviewers registry ──────────────────────────────────────────────────────

@router.get("/interviewers")
def list_interviewers(user: dict = Depends(get_current_user)):
    _require_db()
    return shared_db.get_all_interviewers()


@router.post("/interviewers")
def upsert_interviewer(body: UpsertInterviewerBody, user: dict = Depends(get_current_user)):
    _require_db()
    _require_upload(user)
    iid = shared_db.upsert_interviewer(
        body.interviewer_code, body.name, body.supervisor_id, body.region, body.is_active
    )
    return {"id": iid, **body.model_dump()}


class BulkUpsertSupervisorsBody(BaseModel):
    names: list[str]


@router.post("/supervisors/bulk-upsert")
def bulk_upsert_supervisors(body: BulkUpsertSupervisorsBody, user: dict = Depends(get_current_user)):
    _require_db()
    _require_upload(user)
    return shared_db.upsert_supervisors_bulk(body.names)


@router.get("/interviewers/analytics")
def get_interviewer_analytics(
    project_id: Optional[int] = None,
    supervisor_id: Optional[int] = None,
    region: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    _require_db()
    return shared_db.get_interviewer_analytics(project_id, supervisor_id, region)


@router.get("/interviewers/export")
def export_interviewer_report(
    project_id: Optional[int] = None,
    supervisor_id: Optional[int] = None,
    region: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    """Export a styled Excel report of interviewer performance analytics."""
    import io
    from datetime import datetime
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    _require_db()
    data = shared_db.get_interviewer_analytics(project_id, supervisor_id, region)
    rows = data["interviewers"]
    by_sup = data["by_supervisor"]
    projects = {p["id"]: p["name"] for p in data["projects"]}

    # ── Palette ───────────────────────────────────────────────────────────────
    DARK   = "1A1D2E"
    MID    = "22263A"
    ACCENT = "00C87A"
    RED    = "E84060"
    YELLOW = "E8B931"
    GREEN  = "2ECC87"
    WHITE  = "F0F2FC"
    MUTED  = "8B90A8"

    def fill(hex_str: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_str)

    def hdr_font(bold=True, size=10, color=WHITE) -> Font:
        return Font(name="Calibri", bold=bold, size=size, color=color)

    def cell_font(bold=False, size=10, color="FFFFFF") -> Font:
        return Font(name="Calibri", bold=bold, size=size, color=color)

    thin = Side(style="thin", color="2A2D3E")
    border = Border(bottom=thin)

    def risk_fill(rate: float) -> PatternFill:
        if rate >= 15: return fill(RED)
        if rate >= 8:  return fill(YELLOW)
        return fill(GREEN)

    def risk_font(rate: float) -> Font:
        dark_text = rate >= 8 and rate < 15
        return Font(name="Calibri", bold=True, size=10, color="1A1D2E" if dark_text else WHITE)

    wb = openpyxl.Workbook()

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 1 — Summary / KPIs
    # ═══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22

    def write_kpi(row, label, value, val_color=WHITE):
        ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", size=10, color=MUTED)
        c = ws.cell(row=row, column=2, value=value)
        c.font = Font(name="Calibri", bold=True, size=11, color=val_color)
        ws.row_dimensions[row].height = 20

    # Title block
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "Interviewer Performance Report"
    t.font = Font(name="Calibri", bold=True, size=16, color=ACCENT)
    t.fill = fill(DARK)
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value = f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}   |   Filters: " + \
                ("Project: " + projects.get(project_id, str(project_id)) if project_id else "") + \
                ("  Region: " + region if region else "")
    sub.font = Font(name="Calibri", size=9, color=MUTED)
    sub.fill = fill(DARK)
    ws.row_dimensions[2].height = 16

    ws["A3"].fill = fill(DARK); ws.row_dimensions[3].height = 8

    with_data = [r for r in rows if r["total_interviews"] > 0]
    avg_flag = round(sum(r["flag_rate"] for r in with_data) / len(with_data), 1) if with_data else 0
    high_risk = sum(1 for r in rows if r["flag_rate"] >= 15)
    total_bc_errors = sum(r["bc_errors"] for r in rows)
    total_li_fail = sum(r["li_fail"] for r in rows)
    total_cancelled = sum(r["cancelled"] for r in rows)

    kpis = [
        ("Total Interviewers", len(rows), WHITE),
        ("Active (have records)", len(with_data), ACCENT),
        ("Avg Flag Rate", f"{avg_flag}%", RED if avg_flag >= 15 else YELLOW if avg_flag >= 8 else GREEN),
        ("High Risk (≥15% flag rate)", high_risk, RED if high_risk > 0 else GREEN),
        ("Total BC Errors", total_bc_errors, YELLOW if total_bc_errors > 10 else WHITE),
        ("Total Listen-in Fails", total_li_fail, YELLOW if total_li_fail > 5 else WHITE),
        ("Total Cancelled Interviews", total_cancelled, WHITE),
        ("Supervisors", len(by_sup), WHITE),
    ]
    for i, (label, value, color) in enumerate(kpis, start=4):
        ws.cell(row=i, column=1, value=label).font = Font(name="Calibri", size=10, color=MUTED)
        ws.cell(row=i, column=1).fill = fill(DARK)
        c = ws.cell(row=i, column=2, value=value)
        c.font = Font(name="Calibri", bold=True, size=11, color=color)
        c.fill = fill(DARK)
        ws.row_dimensions[i].height = 22

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 2 — Interviewer Performance
    # ═══════════════════════════════════════════════════════════════════════════
    wi = wb.create_sheet("Interviewer Performance")
    wi.sheet_view.showGridLines = False

    headers = [
        "Code", "Name", "Supervisor", "Region",
        "Interviews", "Dur Flags", "SL Flags", "Flag Rate",
        "BC Sessions", "BC Errors", "Listen-ins", "LI Pass", "LI Fail",
        "Approved", "Cancelled", "Avg Duration (min)", "Risk",
    ]
    col_widths = [12, 22, 20, 14, 12, 12, 10, 11, 12, 11, 12, 10, 10, 12, 12, 18, 8]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        wi.column_dimensions[get_column_letter(col)].width = w
        c = wi.cell(row=1, column=col, value=h)
        c.fill = fill(DARK)
        c.font = hdr_font()
        c.alignment = Alignment(horizontal="center", vertical="center")
    wi.row_dimensions[1].height = 22
    wi.freeze_panes = "A2"

    for r_idx, r in enumerate(rows, start=2):
        rate = r["flag_rate"]
        row_fill = fill(MID) if r_idx % 2 == 0 else fill("1E2135")
        vals = [
            r["code"], r["name"] or "", r["supervisor_name"] or "",
            r["region"] or "", r["total_interviews"],
            r["duration_flags"], r["sl_flags"],
            f"{rate}%" if r["total_interviews"] > 0 else "—",
            r["bc_count"], r["bc_errors"],
            r["li_count"], r["li_pass"], r["li_fail"],
            r["approved"], r["cancelled"],
            r["avg_duration"] if r["total_interviews"] > 0 else "—",
            ("HIGH" if rate >= 15 else "MED" if rate >= 8 else "LOW") if r["total_interviews"] > 0 else "—",
        ]
        for col, val in enumerate(vals, 1):
            c = wi.cell(row=r_idx, column=col, value=val)
            c.fill = row_fill
            c.font = cell_font()
            c.alignment = Alignment(vertical="center")
        # Flag-rate cell (col 8) and Risk cell (col 17) get color treatment
        if r["total_interviews"] > 0:
            for col in (8, 17):
                wi.cell(row=r_idx, column=col).fill = risk_fill(rate)
                wi.cell(row=r_idx, column=col).font = risk_font(rate)
                wi.cell(row=r_idx, column=col).alignment = Alignment(horizontal="center", vertical="center")
        wi.row_dimensions[r_idx].height = 18

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 3 — Supervisor Summary
    # ═══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Supervisor Summary")
    ws3.sheet_view.showGridLines = False

    sup_headers = ["Supervisor", "Interviewers", "Total Interviews", "Total Flags", "Avg Flag Rate", "BC Errors"]
    sup_widths  = [22, 14, 18, 14, 16, 12]
    for col, (h, w) in enumerate(zip(sup_headers, sup_widths), 1):
        ws3.column_dimensions[get_column_letter(col)].width = w
        c = ws3.cell(row=1, column=col, value=h)
        c.fill = fill(DARK); c.font = hdr_font()
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 22
    ws3.freeze_panes = "A2"

    for r_idx, s in enumerate(by_sup, start=2):
        rate = s["flag_rate"]
        row_fill = fill(MID) if r_idx % 2 == 0 else fill("1E2135")
        vals = [s["supervisor_name"], s["interviewer_count"], s["total_interviews"],
                s["total_flags"], f"{rate}%", s["bc_errors"]]
        for col, val in enumerate(vals, 1):
            c = ws3.cell(row=r_idx, column=col, value=val)
            c.fill = row_fill; c.font = cell_font()
            c.alignment = Alignment(vertical="center")
        if s["total_interviews"] > 0:
            for col in (5,):
                ws3.cell(row=r_idx, column=col).fill = risk_fill(rate)
                ws3.cell(row=r_idx, column=col).font = risk_font(rate)
                ws3.cell(row=r_idx, column=col).alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[r_idx].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"interviewer_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/interviewers/{interviewer_code}/metrics")
def get_interviewer_metrics(interviewer_code: str, user: dict = Depends(get_current_user)):
    _require_db()
    return shared_db.get_interviewer_metrics(interviewer_code)


# ── Timing mock-data seed ──────────────────────────────────────────────────────

@router.post("/projects/{project_id}/seed-timing")
def seed_timing(project_id: int, user: dict = Depends(get_current_user)):
    """Generate realistic mock timing records for a project using DB interviewers."""
    _require_db()
    interviewers = shared_db.get_all_interviewers()
    if not interviewers:
        raise HTTPException(404, "No interviewers in DB to seed from")

    rng = random.Random(project_id)
    conn = shared_db.get_conn()
    try:
        start = date(2025, 2, 1)
        records_inserted = 0
        upload_id = str(uuid.uuid4())

        conn.execute(
            """INSERT INTO upload_log (project_id, upload_id, uploaded_by, filename, report_type, row_count, wave_label)
               VALUES (%s, %s, %s, %s, %s, %s, %s)""",
            (project_id, upload_id, user["id"], "demo_timing.xlsx", "timing", 0, "Demo Wave"),
        )

        for itv in interviewers:
            code = itv["interviewer_code"]
            region = itv.get("region") or "Nairobi"
            n = rng.randint(18, 45)
            base_loi = rng.uniform(12, 28)
            for i in range(n):
                day_offset = rng.randint(0, 89)
                interview_date = (start + timedelta(days=day_offset)).isoformat()
                duration = max(3.0, rng.gauss(base_loi, base_loi * 0.18))
                instance_id = f"DEMO-{code}-{i+1:03d}"
                conn.execute(
                    """INSERT INTO timing_records
                       (project_id, upload_id, instance_id, interviewer_id, region, interview_date, duration_minutes)
                       VALUES (%s, %s, %s, %s, %s, %s, %s)""",
                    (project_id, upload_id, instance_id, code, region,
                     interview_date, round(duration, 1)),
                )
                records_inserted += 1

        conn.execute(
            "UPDATE upload_log SET row_count = %s WHERE upload_id = %s",
            (records_inserted, upload_id),
        )
        conn.commit()
    finally:
        conn.close()

    return {"inserted": records_inserted}


# ── Generic filtered Excel export ─────────────────────────────────────────────

NAVY   = "1B2A4A"
TEAL   = "00B5A3"
WHITE  = "FFFFFF"
LIGHT  = "F1F4F8"

def _hdr_fill(hex_col: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_col)

def _thin_border() -> Border:
    s = Side(style="thin", color="D0D5DD")
    return Border(left=s, right=s, top=s, bottom=s)

def _make_chart_image(chart_spec: dict) -> io.BytesIO | None:
    try:
        ctype = chart_spec.get("type", "bar")
        data  = chart_spec.get("data", [])
        title = chart_spec.get("title", "")
        if not data:
            return None

        fig, ax = plt.subplots(figsize=(8, max(3, len(data) * 0.35 + 1)))
        fig.patch.set_facecolor("#F1F4F8")
        ax.set_facecolor("#F1F4F8")
        ax.spines[["top", "right"]].set_visible(False)
        ax.tick_params(colors="#6B7280", labelsize=8)
        for spine in ax.spines.values():
            spine.set_color("#E2E6ED")

        if ctype in ("bar", "bar_horizontal"):
            labels = [str(r.get(chart_spec.get("y", "name"), "")) for r in data]
            values = [float(r.get(chart_spec.get("x", "count"), 0)) for r in data]
            if ctype == "bar_horizontal":
                ax.barh(labels, values, color="#1B2A4A", height=0.6)
                ax.set_xlabel(chart_spec.get("x", "count"), color="#6B7280", fontsize=9)
                ax.invert_yaxis()
            else:
                ax.bar(labels, values, color="#00B5A3", width=0.6)
                ax.set_ylabel(chart_spec.get("y", "count"), color="#6B7280", fontsize=9)
                plt.xticks(rotation=35, ha="right", fontsize=8)

        elif ctype == "bar_stacked":
            series = chart_spec.get("series", [])
            labels = [str(r.get("name", "")) for r in data]
            colors = ["#00B5A3", "#3B5A9A", "#1B2A4A", "#6B7280"]
            bottom = np.zeros(len(data))
            for idx, s in enumerate(series):
                vals = np.array([float(r.get(s, 0)) for r in data])
                ax.barh(labels, vals, left=bottom, color=colors[idx % len(colors)],
                        height=0.6, label=s)
                bottom += vals
            ax.invert_yaxis()
            ax.legend(fontsize=8, loc="lower right")

        ax.set_title(title, color="#1B2A4A", fontsize=11, fontweight="bold", pad=8)
        plt.tight_layout()
        buf = io.BytesIO()
        plt.savefig(buf, format="png", dpi=130, bbox_inches="tight")
        plt.close(fig)
        buf.seek(0)
        return buf
    except Exception:
        return None


class ExportRequest(BaseModel):
    report_type: str
    rows: List[Dict[str, Any]]
    chart_spec: Dict[str, Any] = {}


@router.post("/projects/{project_id}/export")
def export_tab_excel(project_id: int, body: ExportRequest, user: dict = Depends(get_current_user)):
    _require_db()
    rows = body.rows
    chart_spec = body.chart_spec

    wb = Workbook()
    ws = wb.active
    ws.title = body.report_type.replace("_", " ").title()

    hdr_font  = Font(bold=True, color=WHITE, name="Calibri", size=10)
    hdr_fill  = _hdr_fill(NAVY)
    data_fill = _hdr_fill(LIGHT)
    border    = _thin_border()

    if rows:
        cols = list(rows[0].keys())
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col.replace("_", " ").title())
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[cell.column_letter].width = max(12, len(col) + 4)

        for ri, row in enumerate(rows, 2):
            for ci, col in enumerate(cols, 1):
                cell = ws.cell(row=ri, column=ci, value=row.get(col))
                cell.border = border
                cell.alignment = Alignment(vertical="center")
                if ri % 2 == 0:
                    cell.fill = data_fill
        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

    # Chart sheet
    chart_img = _make_chart_image(chart_spec)
    if chart_img:
        wc = wb.create_sheet("Chart")
        wc.sheet_view.showGridLines = False
        img = XLImage(chart_img)
        img.anchor = "B2"
        wc.add_image(img)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from datetime import datetime as _dt
    fname = f"{body.report_type}_report_{project_id}_{_dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/projects/{project_id}/ppt")
async def download_project_ppt(
    project_id: int,
    current_user=Depends(get_current_user),
):
    from ppt_generator import build_project_ppt
    pptx_bytes = build_project_ppt(project_id)
    buf = io.BytesIO(pptx_bytes)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
        headers={"Content-Disposition": f'attachment; filename="QC_Report_Project_{project_id}.pptx"'},
    )
